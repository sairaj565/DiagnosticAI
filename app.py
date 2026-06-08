from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from database.db import init_db, get_connection
from model import predict
import os
import uuid
import json
from flask_mail import Mail, Message
import secrets

from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
load_dotenv()  # loads variables from .env file

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'fallback_dev_key_change_in_production')

app.config['MAIL_SERVER']         = 'smtp.gmail.com'
app.config['MAIL_PORT']           = 587
app.config['MAIL_USE_TLS']        = True
app.config['MAIL_USERNAME']       = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD']       = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_USERNAME')

mail = Mail(app)
reset_tokens = {}

UPLOAD_FOLDER = os.path.join('static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

init_db()


@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, public, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# ==============================
# PACK-YEARS CALCULATOR
# Beedi = 1.5x cigarette multiplier (more toxic, no filter)
# Both  = 1.25x average
# ==============================
def calculate_pack_years(smoking_type, amount_per_day, years_smoked):
    try:
        amount = float(amount_per_day) if amount_per_day else 0.0
        years  = float(years_smoked)   if years_smoked   else 0.0
    except (ValueError, TypeError):
        return 0.0

    if amount <= 0 or years <= 0:
        return 0.0

    if smoking_type == 'Beedi':
        effective_sticks = amount * 1.5
    elif smoking_type == 'Both':
        effective_sticks = amount * 1.25
    else:
        effective_sticks = amount  # Cigarette or Other = 1x

    return round((effective_sticks / 20.0) * years, 1)


@app.route('/')
def landing():
    return render_template('landing.html')


@app.route('/system-status')
def system_status():
    from flask import jsonify
    status = {"db": False, "model": False, "server": True}

    # Check database
    try:
        conn = get_connection()
        conn.execute("SELECT 1 FROM doctors LIMIT 1")
        conn.close()
        status["db"] = True
    except Exception:
        pass

    # Check AI model file exists
    model_paths = [
        os.path.join(os.path.dirname(__file__), 'lung_cancer_model.h5'),
        os.path.join(os.path.dirname(__file__), 'model', 'lung_cancer_model.h5'),
        'lung_cancer_model.h5',
    ]
    status["model"] = any(os.path.exists(p) for p in model_paths)

    if status["db"] and status["model"]:
        status["overall"] = "operational"
        status["message"] = "All Systems Operational"
    elif status["db"] and not status["model"]:
        status["overall"] = "degraded"
        status["message"] = "AI Model Unavailable"
    elif not status["db"] and status["model"]:
        status["overall"] = "degraded"
        status["message"] = "Database Unavailable"
    else:
        status["overall"] = "down"
        status["message"] = "Systems Down"

    return jsonify(status)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email    = request.form.get('email')
        password = request.form.get('password')
        conn     = get_connection()
        cursor   = conn.cursor()
        doctor   = cursor.execute(
            "SELECT * FROM doctors WHERE email = ?", (email,)
        ).fetchone()
        conn.close()
        if doctor and check_password_hash(doctor['password'], password):
            session['doctor']      = email
            session['doctor_name'] = doctor['name']
            session['doctor_spec'] = doctor['specialization'] if doctor['specialization'] else 'Radiologist'
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error="Invalid email or password.")
    return render_template('login.html')


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        name     = request.form.get('name')
        email    = request.form.get('email')
        password = request.form.get('password')
        hospital = request.form.get('hospital')
        spec     = request.form.get('specialization')
        phone    = request.form.get('phone')
        license  = request.form.get('license')

        conn   = get_connection()
        cursor = conn.cursor()
        existing = cursor.execute(
            "SELECT * FROM doctors WHERE email = ?", (email,)
        ).fetchone()
        if existing:
            conn.close()
            return render_template('signup.html', error="Email already registered.")

        cursor.execute('''
            INSERT INTO doctors (name, email, password, hospital, specialization, phone, license)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (name, email, generate_password_hash(password), hospital, spec, phone, license))
        conn.commit()
        conn.close()

        session['doctor']      = email
        session['doctor_name'] = name
        session['doctor_spec'] = spec if spec else 'Radiologist'
        return redirect(url_for('dashboard'))

    return render_template('signup.html')


@app.route('/dashboard')
def dashboard():
    if 'doctor' not in session:
        return redirect(url_for('login'))

    conn   = get_connection()
    cursor = conn.cursor()
    doc    = session['doctor']
    total  = cursor.execute("SELECT COUNT(*) FROM patients WHERE doctor=?", (doc,)).fetchone()[0]
    tumors = cursor.execute("SELECT COUNT(*) FROM patients WHERE doctor=? AND result='Tumor Detected'", (doc,)).fetchone()[0]
    normal = cursor.execute("SELECT COUNT(*) FROM patients WHERE doctor=? AND result='Normal'", (doc,)).fetchone()[0]
    # One row per patient — latest scan only, 5 most recently updated patients
    recent = cursor.execute("""
        SELECT p.*
        FROM patients p
        INNER JOIN (
            SELECT profile_ref_id, MAX(created_at) as latest_at
            FROM patients
            WHERE doctor=? AND profile_ref_id IS NOT NULL AND profile_ref_id != ''
            GROUP BY profile_ref_id
        ) grp ON p.profile_ref_id = grp.profile_ref_id AND p.created_at = grp.latest_at
        WHERE p.doctor=?
        ORDER BY p.created_at DESC
        LIMIT 5
    """, (doc, doc)).fetchall()
    # Visit counts for each profile shown in recent
    visit_counts = {}
    for row in recent:
        ref = row['profile_ref_id']
        if ref:
            count = cursor.execute(
                "SELECT COUNT(*) FROM patients WHERE profile_ref_id=? AND doctor=?", (ref, doc)
            ).fetchone()[0]
            visit_counts[ref] = count

    conn.close()

    return render_template('dashboard.html',
        active_page='dashboard',
        total=total, tumors=tumors, normal=normal, recent=recent,
        visit_counts=visit_counts,
        doctor=session['doctor'],
        doctor_name=session.get('doctor_name', session['doctor'])
    )


@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if 'doctor' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        file           = request.files.get('file')
        name           = request.form.get('name', '').strip()
        age            = request.form.get('age', '').strip()
        gender         = request.form.get('gender', '')
        phone          = request.form.get('phone', '').strip()
        smoking        = request.form.get('smoking', '')
        symptoms_list  = request.form.getlist('symptoms')
        family_history = request.form.get('family_history', '')
        notes          = request.form.get('notes', '').strip()
        scan_type      = request.form.get('scan_type', 'CT Scan')

        # New smoking detail fields
        smoking_type      = request.form.get('smoking_type', '').strip()
        amount_per_day    = request.form.get('amount_per_day', '').strip()
        years_smoked      = request.form.get('years_smoked', '').strip()
        years_since_quit  = request.form.get('years_since_quit', '').strip()

        try:
            amount_per_day_val = float(amount_per_day) if amount_per_day else 0.0
        except ValueError:
            amount_per_day_val = 0.0

        try:
            years_smoked_val = float(years_smoked) if years_smoked else 0.0
        except ValueError:
            years_smoked_val = 0.0

        try:
            years_since_quit_val = float(years_since_quit) if years_since_quit else 0.0
        except ValueError:
            years_since_quit_val = 0.0

        # Calculate pack_years server-side (authoritative)
        if smoking in ('Current Smoker', 'Former Smoker') and smoking_type:
            pack_years_val = calculate_pack_years(smoking_type, amount_per_day_val, years_smoked_val)
        else:
            pack_years_val = 0.0

        symptoms_str = ', '.join(symptoms_list) if symptoms_list else 'None reported'

        clinical_notes_parts = []
        if notes:
            clinical_notes_parts.append(notes)
        if family_history:
            clinical_notes_parts.append(f"Family Hx Cancer: {family_history}")
        clinical_notes = ' | '.join(clinical_notes_parts) if clinical_notes_parts else ''

        patient_type  = request.form.get('patient_type', 'new')
        profile_ref   = request.form.get('profile_ref_id', '').strip()

        patient_id = f"PAT-{uuid.uuid4().hex[:8].upper()}"

        if not name:
            return render_template('upload.html',
                active_page='upload',
                error="Patient name is required.",
                doctor=session['doctor'], doctor_name=session.get('doctor_name', ''))

        if not file or file.filename == '':
            return render_template('upload.html',
                active_page='upload',
                error="Please upload a scan image.",
                doctor=session['doctor'], doctor_name=session.get('doctor_name', ''))

        ext       = os.path.splitext(file.filename)[1].lower()
        filename  = f"{uuid.uuid4().hex}{ext}"
        save_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(save_path)

        try:
            from PIL import Image as PILImage
            img_check     = PILImage.open(save_path)
            width, height = img_check.size
            if width < 224 or height < 224:
                flash(f"Low resolution ({width}×{height}px) — minimum 224×224 recommended.", "warning")
            elif width * height < 100000:
                flash(f"Image resolution ({width}×{height}) may miss fine margins.", "warning")
        except Exception as e:
            print(f"Resolution check failed: {e}")

        result = predict(save_path, age=age, smoking=smoking)

        if 'error' in result:
            return render_template('upload.html',
                active_page='upload',
                error=result['error'],
                doctor=session['doctor'], doctor_name=session.get('doctor_name', ''))

        conn   = get_connection()
        cursor = conn.cursor()

        # ── Create or find patient profile ──────────────────────────────
        if patient_type == 'returning' and profile_ref:
            # Link to existing profile — just verify it belongs to this doctor
            existing_profile = cursor.execute(
                "SELECT ref_id FROM patient_profiles WHERE ref_id=? AND doctor=?",
                (profile_ref, session['doctor'])
            ).fetchone()
            if existing_profile:
                final_ref_id = profile_ref
            else:
                # Profile not found — create new one as fallback
                final_ref_id = f"REF-{uuid.uuid4().hex[:8].upper()}"
                cursor.execute(
                    "INSERT INTO patient_profiles (ref_id, name, gender, phone, doctor) VALUES (?,?,?,?,?)",
                    (final_ref_id, name, gender, phone, session['doctor'])
                )
        else:
            # New patient — create new profile
            final_ref_id = f"REF-{uuid.uuid4().hex[:8].upper()}"
            cursor.execute(
                "INSERT INTO patient_profiles (ref_id, name, gender, phone, doctor) VALUES (?,?,?,?,?)",
                (final_ref_id, name, gender, phone, session['doctor'])
            )

        cursor.execute('''
            INSERT INTO patients
            (patient_id, name, age, gender, smoking, phone, notes,
             scan_type, image_path, result, confidence, doctor,
             pack_years, symptoms,
             smoking_type, amount_per_day, years_smoked, years_since_quit,
             profile_ref_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            patient_id, name, age, gender, smoking, phone, clinical_notes,
            scan_type, filename, result['result'], result['confidence'],
            session['doctor'],
            pack_years_val, symptoms_str,
            smoking_type, amount_per_day_val, years_smoked_val, years_since_quit_val,
            final_ref_id
        ))
        conn.commit()
        conn.close()

        loc      = result.get('location', {})
        raw_lobe = loc.get('lobe', '—')
        side     = loc.get('side', '—')

        if 'Left' in side and raw_lobe == 'Middle Lobe':
            final_lobe = 'Left Upper Lobe (Lingula)'
        elif 'Right' in side and raw_lobe == 'Middle Lobe':
            final_lobe = 'Right Middle Lobe'
        else:
            final_lobe = raw_lobe

        session['last_result'] = {
            "patient_id":       patient_id,
            "name":             name,
            "age":              int(age) if age else 0,
            "gender":           gender,
            "smoking":          smoking,
            "smoking_type":     smoking_type,
            "amount_per_day":   amount_per_day_val,
            "years_smoked":     years_smoked_val,
            "years_since_quit": years_since_quit_val,
            "pack_years":       pack_years_val,
            "phone":            phone,
            "symptoms":         symptoms_str,
            "family_history":   family_history,
            "notes":            notes,
            "scan_type":        scan_type,
            "image":            filename,
            "result":           result['result'],
            "disease":          result['disease'],
            "confidence":       result['confidence'],
            "risk":             result['risk'],
            "color":            result['color'],
            "raw_pred":         result.get('raw_pred', 0),
            "heatmap_img":      result.get('heatmap_img', ''),
            "bbox_img":         result.get('bbox_img', ''),
            "size_mm":          result.get('size_mm', 0),
            "size_label":       result.get('size_label', 'N/A'),
            "lungrads":         result.get('lungrads', {}),
            "cancer_type":      result.get('cancer_type', None),
            "location": {
                "anatomical":    loc.get('anatomical', '—'),
                "side":          side,
                "lobe":          final_lobe,
                "left_pct":      loc.get('left_pct', 0),
                "right_pct":     loc.get('right_pct', 0),
                "top_pct":       loc.get('top_pct', 0),
                "mid_pct":       loc.get('mid_pct', 0),
                "bot_pct":       loc.get('bot_pct', 0),
                "dominant_side": loc.get('dominant_side', '—'),
                "dominant_pct":  loc.get('dominant_pct', 0),
            },
        }

        return redirect(url_for('result'))

    return render_template('upload.html',
        active_page='upload',
        doctor=session['doctor'],
        doctor_name=session.get('doctor_name', ''))


@app.route('/result')
def result():
    if 'doctor' not in session:
        return redirect(url_for('login'))
    data = session.get('last_result')
    if not data:
        return redirect(url_for('upload'))
    return render_template('result.html',
        active_page='upload',
        data=data, doctor=session['doctor'],
        doctor_name=session.get('doctor_name', ''))


@app.route('/result/<patient_id>')
def result_by_id(patient_id):
    if 'doctor' not in session:
        return redirect(url_for('login'))

    conn   = get_connection()
    cursor = conn.cursor()
    p      = cursor.execute(
        "SELECT * FROM patients WHERE patient_id = ?", (patient_id,)
    ).fetchone()
    conn.close()

    if not p:
        flash(f"Patient record '{patient_id}' not found.", "error")
        return redirect(url_for('records'))

    p = dict(p)

    notes_raw      = p.get('notes') or ''
    family_history = ''
    notes_clean    = ''

    for part in notes_raw.split(' | '):
        if part.startswith('Family Hx Cancer:'):
            family_history = part.replace('Family Hx Cancer:', '').strip()
        else:
            if part.strip():
                notes_clean = part.strip()

    data = {
        "patient_id":       p.get('patient_id', '—'),
        "name":             p.get('name') or '—',
        "age":              p.get('age') or 0,
        "gender":           p.get('gender') or '—',
        "smoking":          p.get('smoking') or '—',
        "smoking_type":     p.get('smoking_type') or '',
        "amount_per_day":   p.get('amount_per_day') or 0.0,
        "years_smoked":     p.get('years_smoked') or 0.0,
        "years_since_quit": p.get('years_since_quit') or 0.0,
        "pack_years":       p.get('pack_years') or 0.0,
        "phone":            p.get('phone') or '—',
        "symptoms":         p.get('symptoms') or 'None reported',
        "family_history":   family_history or '—',
        "notes":            notes_clean,
        "scan_type":        p.get('scan_type') or 'CT Scan',
        "image":            p.get('image_path') or '',
        "result":           p.get('result', '—'),
        "disease":          p.get('result', '—'),
        "confidence":       p.get('confidence', 0),
        "risk":             "HIGH" if p.get('result') == 'Tumor Detected' else "LOW",
        "color":            "#ef4444" if p.get('result') == 'Tumor Detected' else "#10b981",
        "raw_pred":         float(p.get('confidence', 0)) / 100 if p.get('confidence') else 0,
        "heatmap_img":      "",
        "bbox_img":         "",
        "size_mm":          0,
        "size_label":       "N/A",
        "cancer_type":      None,
        "lungrads": {
            "category":    "4B" if p.get('result') == 'Tumor Detected' else "1",
            "label":       "Very Suspicious" if p.get('result') == 'Tumor Detected' else "Negative",
            "description": "Highly suspicious for malignancy. Biopsy or tissue sampling is indicated." if p.get('result') == 'Tumor Detected' else "No lung nodules present. Routine annual screening recommended.",
            "action":      "Tissue sampling recommended" if p.get('result') == 'Tumor Detected' else "Continue annual screening",
            "color":       "#ef4444" if p.get('result') == 'Tumor Detected' else "#10b981",
            "badge":       "#ef4444" if p.get('result') == 'Tumor Detected' else "#10b981",
        },
        "location": {
            "anatomical": "—", "side": "—", "lobe": "—",
            "left_pct": 0, "right_pct": 0, "top_pct": 0,
            "mid_pct": 0, "bot_pct": 0, "dominant_side": "—", "dominant_pct": 0,
        },
    }

    return render_template('result.html',
        active_page='records',
        data=data, doctor=session['doctor'],
        doctor_name=session.get('doctor_name', ''))


# ── RECORDS — one row per patient profile (no duplicates) ──────
@app.route('/records')
def records():
    if 'doctor' not in session:
        return redirect(url_for('login'))

    conn = get_connection()
    doc  = session['doctor']

    # ── Step 1: Fix orphan scans (scans with no profile_ref_id) ──
    # Groups by name so the same patient gets ONE profile, not one per scan
    orphan_names = conn.execute(
        """SELECT DISTINCT LOWER(TRIM(name)) as norm_name, name, gender, phone
           FROM patients
           WHERE doctor=? AND (profile_ref_id IS NULL OR profile_ref_id='')
           ORDER BY name""",
        (doc,)
    ).fetchall()

    for row in orphan_names:
        existing = conn.execute(
            "SELECT ref_id FROM patient_profiles WHERE doctor=? AND LOWER(TRIM(name))=?",
            (doc, row['norm_name'])
        ).fetchone()

        if existing:
            ref_id = existing['ref_id']
        else:
            ref_id = f"REF-{uuid.uuid4().hex[:8].upper()}"
            conn.execute(
                "INSERT INTO patient_profiles (ref_id, name, gender, phone, doctor) VALUES (?,?,?,?,?)",
                (ref_id, row['name'], row['gender'], row['phone'], doc)
            )

        conn.execute(
            """UPDATE patients SET profile_ref_id=?
               WHERE doctor=? AND LOWER(TRIM(name))=?
               AND (profile_ref_id IS NULL OR profile_ref_id='')""",
            (ref_id, doc, row['norm_name'])
        )

    if orphan_names:
        conn.commit()

    # ── Step 2: Merge duplicate profiles (same name → keep oldest) ──
    dup_names = conn.execute(
        """SELECT LOWER(TRIM(name)) as norm_name
           FROM patient_profiles WHERE doctor=?
           GROUP BY LOWER(TRIM(name)) HAVING COUNT(*) > 1""",
        (doc,)
    ).fetchall()

    for dup in dup_names:
        # Keep the profile with the MOST scans; if tied, keep oldest rowid
        profiles = conn.execute(
            """SELECT pp.ref_id, COUNT(p.patient_id) as scan_count
               FROM patient_profiles pp
               LEFT JOIN patients p ON p.profile_ref_id=pp.ref_id AND p.doctor=pp.doctor
               WHERE pp.doctor=? AND LOWER(TRIM(pp.name))=?
               GROUP BY pp.ref_id
               ORDER BY scan_count DESC, pp.rowid ASC""",
            (doc, dup['norm_name'])
        ).fetchall()

        keep_ref  = profiles[0]['ref_id']
        dupe_refs = [p['ref_id'] for p in profiles[1:]]

        for dupe_ref in dupe_refs:
            conn.execute(
                "UPDATE patients SET profile_ref_id=? WHERE profile_ref_id=? AND doctor=?",
                (keep_ref, dupe_ref, doc)
            )
            conn.execute(
                "DELETE FROM patient_profiles WHERE ref_id=? AND doctor=?",
                (dupe_ref, doc)
            )

    if dup_names:
        conn.commit()

    # ── Step 3: Delete ghost profiles (profile exists but zero scans) ──
    conn.execute(
        """DELETE FROM patient_profiles WHERE doctor=?
           AND ref_id NOT IN (
               SELECT DISTINCT profile_ref_id FROM patients
               WHERE doctor=? AND profile_ref_id IS NOT NULL AND profile_ref_id != ''
           )""",
        (doc, doc)
    )
    conn.commit()

    # ── Step 4: Fetch one row per profile with latest scan info ──
    patients = conn.execute("""
        SELECT
            pp.ref_id           AS profile_ref_id,
            pp.name,
            pp.gender,
            pp.phone,
            COUNT(p.patient_id) AS visit_count,
            MAX(p.created_at)   AS last_visit,
            (SELECT patient_id FROM patients WHERE profile_ref_id=pp.ref_id AND doctor=pp.doctor ORDER BY created_at DESC LIMIT 1) AS patient_id,
            (SELECT age        FROM patients WHERE profile_ref_id=pp.ref_id AND doctor=pp.doctor ORDER BY created_at DESC LIMIT 1) AS age,
            (SELECT result     FROM patients WHERE profile_ref_id=pp.ref_id AND doctor=pp.doctor ORDER BY created_at DESC LIMIT 1) AS result,
            (SELECT confidence FROM patients WHERE profile_ref_id=pp.ref_id AND doctor=pp.doctor ORDER BY created_at DESC LIMIT 1) AS confidence,
            (SELECT smoking    FROM patients WHERE profile_ref_id=pp.ref_id AND doctor=pp.doctor ORDER BY created_at DESC LIMIT 1) AS smoking,
            (SELECT scan_type  FROM patients WHERE profile_ref_id=pp.ref_id AND doctor=pp.doctor ORDER BY created_at DESC LIMIT 1) AS scan_type,
            (SELECT image_path FROM patients WHERE profile_ref_id=pp.ref_id AND doctor=pp.doctor ORDER BY created_at DESC LIMIT 1) AS image_path,
            (SELECT created_at FROM patients WHERE profile_ref_id=pp.ref_id AND doctor=pp.doctor ORDER BY created_at DESC LIMIT 1) AS created_at
        FROM patient_profiles pp
        LEFT JOIN patients p ON p.profile_ref_id=pp.ref_id AND p.doctor=pp.doctor
        WHERE pp.doctor=?
        GROUP BY pp.ref_id
        ORDER BY last_visit DESC
    """, (doc,)).fetchall()

    total_scans = conn.execute("SELECT COUNT(*) FROM patients WHERE doctor=?", (doc,)).fetchone()[0]
    conn.close()
    return render_template('records.html',
        active_page='records',
        patients=patients,
        total_scans=total_scans,
        doctor=session['doctor'],
        doctor_name=session.get('doctor_name', session['doctor'])
    )


# ── Delete patient ─────────────────────────────────────────────
@app.route('/patient/delete/<patient_id>', methods=['POST'])
def delete_patient(patient_id):
    if 'doctor' not in session:
        return redirect(url_for('login'))
    conn = get_connection()
    # Only allow deleting own patients
    p = conn.execute(
        "SELECT * FROM patients WHERE patient_id=? AND doctor=?",
        (patient_id, session['doctor'])
    ).fetchone()
    if p:
        # Delete scan image file
        if p['image_path']:
            img_path = os.path.join(UPLOAD_FOLDER, p['image_path'])
            if os.path.exists(img_path):
                os.remove(img_path)
        conn.execute("DELETE FROM patients WHERE patient_id=?", (patient_id,))
        conn.commit()
    conn.close()
    flash(f"Patient record {patient_id} deleted.", "success")
    return redirect(url_for('records'))


# ── Edit patient ───────────────────────────────────────────────
@app.route('/patient/edit/<patient_id>', methods=['GET', 'POST'])
def edit_patient(patient_id):
    if 'doctor' not in session:
        return redirect(url_for('login'))

    conn = get_connection()
    p = conn.execute(
        "SELECT * FROM patients WHERE patient_id=? AND doctor=?",
        (patient_id, session['doctor'])
    ).fetchone()

    if not p:
        conn.close()
        flash("Patient not found.", "error")
        return redirect(url_for('records'))

    if request.method == 'POST':
        name             = request.form.get('name', '').strip()
        age              = request.form.get('age', '').strip()
        gender           = request.form.get('gender', '')
        phone            = request.form.get('phone', '').strip()
        smoking          = request.form.get('smoking', '')
        smoking_type     = request.form.get('smoking_type', '').strip()
        amount_per_day   = request.form.get('amount_per_day', '').strip()
        years_smoked     = request.form.get('years_smoked', '').strip()
        years_since_quit = request.form.get('years_since_quit', '').strip()
        symptoms_list    = request.form.getlist('symptoms')
        notes            = request.form.get('notes', '').strip()
        family_history   = request.form.get('family_history', '')
        scan_type        = request.form.get('scan_type', 'CT Scan')

        try: amount_per_day_val = float(amount_per_day) if amount_per_day else 0.0
        except ValueError: amount_per_day_val = 0.0
        try: years_smoked_val = float(years_smoked) if years_smoked else 0.0
        except ValueError: years_smoked_val = 0.0
        try: years_since_quit_val = float(years_since_quit) if years_since_quit else 0.0
        except ValueError: years_since_quit_val = 0.0

        if smoking in ('Current Smoker', 'Former Smoker') and smoking_type:
            pack_years_val = calculate_pack_years(smoking_type, amount_per_day_val, years_smoked_val)
        else:
            pack_years_val = 0.0

        symptoms_str = ', '.join(symptoms_list) if symptoms_list else 'None reported'
        clinical_notes_parts = []
        if notes: clinical_notes_parts.append(notes)
        if family_history: clinical_notes_parts.append(f"Family Hx Cancer: {family_history}")
        clinical_notes = ' | '.join(clinical_notes_parts) if clinical_notes_parts else ''

        conn.execute('''
            UPDATE patients SET
              name=?, age=?, gender=?, phone=?, smoking=?,
              smoking_type=?, amount_per_day=?, years_smoked=?, years_since_quit=?,
              pack_years=?, symptoms=?, notes=?, scan_type=?
            WHERE patient_id=? AND doctor=?
        ''', (
            name, age, gender, phone, smoking,
            smoking_type, amount_per_day_val, years_smoked_val, years_since_quit_val,
            pack_years_val, symptoms_str, clinical_notes, scan_type,
            patient_id, session['doctor']
        ))
        conn.commit()
        conn.close()
        flash(f"Patient {name} updated successfully.", "success")
        return redirect(url_for('result_by_id', patient_id=patient_id))

    conn.close()
    # Parse notes and family history for pre-filling
    notes_raw = p['notes'] or ''
    family_history = ''
    notes_clean = ''
    for part in notes_raw.split(' | '):
        if part.startswith('Family Hx Cancer:'):
            family_history = part.replace('Family Hx Cancer:', '').strip()
        elif part.strip():
            notes_clean = part.strip()

    symptoms_list = [s.strip() for s in (p['symptoms'] or '').split(',') if s.strip() and s.strip() != 'None reported']

    return render_template('edit_patient.html',
        active_page='records',
        p=p, patient_id=patient_id,
        notes_clean=notes_clean,
        family_history=family_history,
        symptoms_list=symptoms_list,
        doctor=session['doctor'],
        doctor_name=session.get('doctor_name', '')
    )


@app.route('/reports')
def reports():
    if 'doctor' not in session:
        return redirect(url_for('login'))
    conn     = get_connection()
    cursor   = conn.cursor()
    patients = cursor.execute(
        "SELECT * FROM patients WHERE doctor=? ORDER BY created_at DESC",
        (session['doctor'],)
    ).fetchall()
    profile = cursor.execute(
        "SELECT * FROM doctors WHERE email=?", (session['doctor'],)
    ).fetchone()
    conn.close()
    return render_template('reports.html',
        active_page='reports',
        patients=patients, doctor=session['doctor'],
        doctor_name=session.get('doctor_name', session['doctor']),
        profile=profile)





@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('landing'))


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email  = request.form.get('email')
        conn   = get_connection()
        cursor = conn.cursor()
        doctor = cursor.execute(
            "SELECT * FROM doctors WHERE email = ?", (email,)
        ).fetchone()
        conn.close()

        if not doctor:
            return render_template('forgot_password.html',
                error="No account found with this email address.")

        token  = secrets.token_urlsafe(32)
        conn   = get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM reset_tokens WHERE email = ?", (email,))
        cursor.execute("INSERT INTO reset_tokens (email, token) VALUES (?, ?)", (email, token))
        conn.commit()
        conn.close()

        # Redirect directly to the local Play-Test verification challenge
        return redirect(url_for('forgot_password_challenge', token=token))

    return render_template('forgot_password.html')


@app.route('/forgot-password/challenge/<token>', methods=['GET', 'POST'])
def forgot_password_challenge(token):
    conn   = get_connection()
    record = conn.execute("SELECT * FROM reset_tokens WHERE token = ?", (token,)).fetchone()
    conn.close()

    if not record:
        return render_template('forgot_password.html', error="Invalid or expired reset session.")

    if request.method == 'POST':
        answer = request.form.get('answer')
        if answer == 'Lungs':
            # Correct answer! Redirect directly to reset-password
            return redirect(url_for('reset_password', token=token))
        else:
            return render_template('forgot_password_challenge.html',
                token=token, email=record['email'], error="Incorrect selection. Please identify the correct organ.")

    return render_template('forgot_password_challenge.html', token=token, email=record['email'])


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    conn   = get_connection()
    record = conn.execute("SELECT * FROM reset_tokens WHERE token = ?", (token,)).fetchone()
    conn.close()

    if not record:
        return render_template('reset_password.html', error="Invalid or expired reset link.")

    if request.method == 'POST':
        new_pass = request.form.get('new_password')
        confirm  = request.form.get('confirm_password')
        if new_pass != confirm:
            return render_template('reset_password.html', token=token, error="Passwords do not match.")
        if len(new_pass) < 6:
            return render_template('reset_password.html', token=token, error="Password must be at least 6 characters.")

        conn = get_connection()
        conn.execute("UPDATE doctors SET password = ? WHERE email = ?", (generate_password_hash(new_pass), record['email']))
        conn.execute("DELETE FROM reset_tokens WHERE token = ?", (token,))
        conn.commit()
        conn.close()
        return render_template('reset_password.html', success=True)

    return render_template('reset_password.html', token=token)


@app.route('/settings', methods=['GET'])
def settings():
    if 'doctor' not in session:
        return redirect(url_for('login'))
    if not session.get('settings_verified'):
        return redirect(url_for('settings_verify'))

    conn   = get_connection()
    doctor = conn.execute("SELECT * FROM doctors WHERE email = ?", (session['doctor'],)).fetchone()
    conn.close()

    return render_template('settings.html',
        active_page='settings',
        doctor=session['doctor'], doctor_name=session.get('doctor_name', ''),
        profile=doctor, success=None, error=None)


@app.route('/settings/verify', methods=['GET', 'POST'])
def settings_verify():
    if 'doctor' not in session:
        return redirect(url_for('login'))

    attempts = session.get('settings_attempts', 0)
    locked   = attempts >= 3

    if request.method == 'POST':
        if locked:
            return render_template('settings_lock.html',
                doctor=session['doctor'], doctor_name=session.get('doctor_name', ''),
                error="Too many attempts. Try again later.", locked=True, attempts=attempts)

        password = request.form.get('password')
        conn   = get_connection()
        doctor = conn.execute(
            "SELECT * FROM doctors WHERE email = ?",
            (session['doctor'],)
        ).fetchone()
        conn.close()

        if doctor and check_password_hash(doctor['password'], password):
            session['settings_verified'] = True
            session['settings_attempts'] = 0
            return redirect(url_for('settings'))
        else:
            session['settings_attempts'] = attempts + 1
            remaining = 3 - session['settings_attempts']
            error_msg = f"Incorrect password. {remaining} attempt{'s' if remaining != 1 else ''} remaining."
            if session['settings_attempts'] >= 3:
                error_msg = "Account locked. Too many failed attempts."
            return render_template('settings_lock.html',
                doctor=session['doctor'], doctor_name=session.get('doctor_name', ''),
                error=error_msg, locked=session['settings_attempts'] >= 3,
                attempts=session['settings_attempts'])

    return render_template('settings_lock.html',
        doctor=session['doctor'], doctor_name=session.get('doctor_name', ''),
        error=None, locked=locked, attempts=attempts)


@app.route('/settings/profile', methods=['POST'])
def settings_profile():
    if 'doctor' not in session:
        return redirect(url_for('login'))

    name     = request.form.get('name')
    hospital = request.form.get('hospital')
    spec     = request.form.get('specialization')
    phone    = request.form.get('phone')

    conn = get_connection()
    conn.execute(
        "UPDATE doctors SET name=?, hospital=?, specialization=?, phone=? WHERE email=?",
        (name, hospital, spec, phone, session['doctor'])
    )
    conn.commit()
    doctor = conn.execute("SELECT * FROM doctors WHERE email = ?", (session['doctor'],)).fetchone()
    conn.close()

    session['doctor_name'] = name
    session['doctor_spec'] = spec

    return render_template('settings.html',
        active_page='settings',
        doctor=session['doctor'], doctor_name=session.get('doctor_name', ''),
        profile=doctor, success='profile', error=None)


@app.route('/settings/password', methods=['POST'])
def settings_password():
    if 'doctor' not in session:
        return redirect(url_for('login'))

    current  = request.form.get('current_password')
    new_pass = request.form.get('new_password')
    confirm  = request.form.get('confirm_password')

    conn   = get_connection()
    doctor = conn.execute(
        "SELECT * FROM doctors WHERE email = ?",
        (session['doctor'],)
    ).fetchone()

    if not doctor or not check_password_hash(doctor['password'], current):
        profile = conn.execute("SELECT * FROM doctors WHERE email = ?", (session['doctor'],)).fetchone()
        conn.close()
        return render_template('settings.html',
            active_page='settings',
            doctor=session['doctor'], doctor_name=session.get('doctor_name', ''),
            profile=profile, success=None, error='wrong_password')

    if new_pass != confirm:
        conn.close()
        return render_template('settings.html',
            active_page='settings',
            doctor=session['doctor'], doctor_name=session.get('doctor_name', ''),
            profile=doctor, success=None, error='password_mismatch')

    conn.execute("UPDATE doctors SET password=? WHERE email=?", (generate_password_hash(new_pass), session['doctor']))
    conn.commit()
    conn.close()

    return render_template('settings.html',
        active_page='settings',
        doctor=session['doctor'], doctor_name=session.get('doctor_name', ''),
        profile=doctor, success='password', error=None)


@app.route('/settings/reset-lock')
def settings_reset_lock():
    if 'doctor' not in session:
        return redirect(url_for('login'))
    session['settings_attempts'] = 0
    session['settings_verified'] = False
    return redirect(url_for('settings_verify'))


@app.route('/export/excel')
def export_excel():
    if 'doctor' not in session:
        return redirect(url_for('login'))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import make_response
    import io

    conn = get_connection()
    patients = conn.execute(
        "SELECT * FROM patients WHERE doctor=? ORDER BY created_at DESC",
        (session['doctor'],)
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patient Records"

    # Header style
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center")
    border      = Border(
        bottom=Side(style="thin", color="3B82F6"),
        top=Side(style="thin", color="3B82F6")
    )

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value     = f"DiagnosticAI — Patient Records | Dr. {session.get('doctor_name','')} | Exported {__import__('datetime').datetime.now().strftime('%d %b %Y')}"
    title_cell.font      = Font(bold=True, color="60A5FA", size=13)
    title_cell.fill      = PatternFill("solid", fgColor="0A0F1E")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = ["Patient ID", "Name", "Age", "Gender", "Scan Type", "Result", "Confidence (%)", "Smoking", "Pack Years", "Date"]
    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=2, column=col, value=h)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center
        cell.border     = border
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, p in enumerate(patients, 3):
        row_fill = PatternFill("solid", fgColor="111827") if row_idx % 2 == 0 else PatternFill("solid", fgColor="0D1525")
        values = [
            p['patient_id'],
            p['name'] or '—',
            p['age'] or '—',
            p['gender'] or '—',
            p['scan_type'] or '—',
            p['result'] or '—',
            p['confidence'] or 0,
            p['smoking'] or '—',
            p['pack_years'] or 0,
            str(p['created_at'])[:10] if p['created_at'] else '—'
        ]
        for col, val in enumerate(values, 1):
            cell           = ws.cell(row=row_idx, column=col, value=val)
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(color="F0F4FF", size=10)
            if col == 6:
                if val == 'Tumor Detected':
                    cell.font = Font(color="F87171", bold=True, size=10)
                elif val == 'Normal':
                    cell.font = Font(color="34D399", bold=True, size=10)
            if col == 7 and p['result'] == 'Tumor Detected':
                cell.font = Font(color="F87171", size=10)
            elif col == 7:
                cell.font = Font(color="34D399", size=10)
        ws.row_dimensions[row_idx].height = 20

    col_widths = [18, 20, 8, 12, 14, 18, 16, 18, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname    = f"DiagnosticAI_Records_{session.get('doctor_name','').replace(' ','_')}_{__import__('datetime').datetime.now().strftime('%Y%m%d')}.xlsx"
    response = make_response(buf.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    response.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export/pdf')
def export_pdf():
    if 'doctor' not in session:
        return redirect(url_for('login'))
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from flask import make_response
    import io
    from datetime import datetime

    conn = get_connection()
    patients = conn.execute(
        "SELECT * FROM patients WHERE doctor=? ORDER BY created_at DESC",
        (session['doctor'],)
    ).fetchall()
    profile = conn.execute(
        "SELECT * FROM doctors WHERE email=?", (session['doctor'],)
    ).fetchone()
    conn.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm)

    navy    = colors.HexColor("#0A0F1E")
    blue    = colors.HexColor("#2563EB")
    blue_l  = colors.HexColor("#3B82F6")
    red     = colors.HexColor("#EF4444")
    green   = colors.HexColor("#10B981")
    muted   = colors.HexColor("#94A3B8")
    white   = colors.white
    dark    = colors.HexColor("#111827")
    mid     = colors.HexColor("#1E2D45")

    styles  = getSampleStyleSheet()
    title_s = ParagraphStyle('title', fontSize=18, textColor=white,    fontName='Helvetica-Bold', spaceAfter=4)
    sub_s   = ParagraphStyle('sub',   fontSize=10, textColor=muted,    fontName='Helvetica',      spaceAfter=2)
    body_s  = ParagraphStyle('body',  fontSize=9,  textColor=white,    fontName='Helvetica')
    small_s = ParagraphStyle('small', fontSize=7,  textColor=muted,    fontName='Helvetica')

    elems = []

    elems.append(Paragraph("🫁 DiagnosticAI", title_s))
    elems.append(Paragraph("AI-Powered Lung Cancer Detection Platform", sub_s))
    elems.append(Paragraph(
        f"Dr. {session.get('doctor_name','')}  |  {profile['hospital'] if profile and profile['hospital'] else ''}  |  Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
        sub_s))
    elems.append(Spacer(1, 8*mm))

    total  = len(patients)
    tumors = sum(1 for p in patients if p['result'] == 'Tumor Detected')
    normal = total - tumors

    summary_data = [
        [Paragraph(f"<b><font color='#60A5FA' size=16>{total}</font></b><br/><font color='#94A3B8' size=8>Total Patients</font>", body_s),
         Paragraph(f"<b><font color='#F87171' size=16>{tumors}</font></b><br/><font color='#94A3B8' size=8>Tumor Detected</font>", body_s),
         Paragraph(f"<b><font color='#34D399' size=16>{normal}</font></b><br/><font color='#94A3B8' size=8>Normal / Healthy</font>", body_s),
         Paragraph(f"<b><font color='#FBBF24' size=16>99.4%</font></b><br/><font color='#94A3B8' size=8>Model Accuracy</font>", body_s)]
    ]
    summary_tbl = Table(summary_data, colWidths=[44*mm]*4)
    summary_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), dark),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [dark]),
        ('BOX',         (0,0), (-1,-1), 0.5, blue_l),
        ('INNERGRID',   (0,0), (-1,-1), 0.3, mid),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',  (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
    ]))
    elems.append(summary_tbl)
    elems.append(Spacer(1, 6*mm))

    elems.append(Paragraph("Patient Records", ParagraphStyle('h2', fontSize=12, textColor=blue_l, fontName='Helvetica-Bold', spaceAfter=4)))

    col_headers = ["Patient ID", "Name", "Age", "Result", "Conf %", "Smoking", "Scan Type", "Date"]
    table_data  = [col_headers]

    for p in patients:
        result_txt = p['result'] or '—'
        table_data.append([
            str(p['patient_id'] or '—'),
            str(p['name'] or '—'),
            str(p['age'] or '—'),
            result_txt,
            f"{p['confidence'] or 0}%",
            str(p['smoking'] or '—'),
            str(p['scan_type'] or '—'),
            str(p['created_at'])[:10] if p['created_at'] else '—'
        ])

    col_w = [30*mm, 32*mm, 12*mm, 30*mm, 16*mm, 28*mm, 20*mm, 22*mm]
    tbl   = Table(table_data, colWidths=col_w, repeatRows=1)

    tbl_styles = [
        ('BACKGROUND',    (0,0), (-1,0),  blue),
        ('TEXTCOLOR',     (0,0), (-1,0),  white),
        ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0),  8),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE',      (0,1), (-1,-1), 8),
        ('TEXTCOLOR',     (0,1), (-1,-1), white),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [dark, mid]),
        ('BOX',           (0,0), (-1,-1), 0.5, blue_l),
        ('INNERGRID',     (0,0), (-1,-1), 0.2, colors.HexColor("#1E2D45")),
    ]
    for i, p in enumerate(patients, 1):
        c = red if p['result'] == 'Tumor Detected' else green
        tbl_styles.append(('TEXTCOLOR', (3,i), (3,i), c))
        tbl_styles.append(('FONTNAME',  (3,i), (3,i), 'Helvetica-Bold'))

    tbl.setStyle(TableStyle(tbl_styles))
    elems.append(tbl)

    elems.append(Spacer(1, 8*mm))
    elems.append(Paragraph(
        "This report was generated by DiagnosticAI. AI results are for clinical decision support only and must be confirmed by a qualified radiologist.",
        small_s))

    doc.build(elems)
    buf.seek(0)

    fname    = f"DiagnosticAI_Report_{session.get('doctor_name','').replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    response = make_response(buf.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    response.headers['Content-Type']        = 'application/pdf'
    return response


@app.route('/patient/search')
def patient_search():
    from flask import jsonify
    if 'doctor' not in session:
        return jsonify([])
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    conn = get_connection()
    profiles = conn.execute(
        """SELECT ref_id, name, gender, phone,
           (SELECT COUNT(*) FROM patients WHERE profile_ref_id=patient_profiles.ref_id) as visit_count,
           (SELECT created_at FROM patients WHERE profile_ref_id=patient_profiles.ref_id ORDER BY created_at DESC LIMIT 1) as last_visit
           FROM patient_profiles
           WHERE doctor=? AND (name LIKE ? OR phone LIKE ?)
           ORDER BY name ASC LIMIT 8""",
        (session['doctor'], f'%{q}%', f'%{q}%')
    ).fetchall()
    conn.close()
    return jsonify([dict(p) for p in profiles])


@app.route('/patient/profile/<ref_id>')
def patient_profile(ref_id):
    if 'doctor' not in session:
        return redirect(url_for('login'))
    conn = get_connection()
    profile = conn.execute(
        "SELECT * FROM patient_profiles WHERE ref_id=? AND doctor=?",
        (ref_id, session['doctor'])
    ).fetchone()
    if not profile:
        flash("Patient profile not found.", "error")
        return redirect(url_for('records'))
    scans = conn.execute(
        "SELECT * FROM patients WHERE profile_ref_id=? AND doctor=? ORDER BY created_at DESC",
        (ref_id, session['doctor'])
    ).fetchall()
    conn.close()
    return render_template('patient_profile.html',
        active_page='records',
        profile=profile, scans=scans,
        doctor=session['doctor'],
        doctor_name=session.get('doctor_name', ''),
        ref_id=ref_id)


@app.route('/analytics')
def analytics():
    if 'doctor' not in session:
        return redirect(url_for('login'))

    conn = get_connection()
    rows = conn.execute(
        "SELECT patient_id, name, age, gender, smoking, scan_type, result, confidence, created_at "
        "FROM patients WHERE doctor = ? ORDER BY created_at ASC",
        (session['doctor'],)
    ).fetchall()
    conn.close()

    def derive_risk(result, confidence):
        if result == 'Tumor Detected':
            return 'HIGH' if confidence >= 85 else 'MEDIUM'
        return 'LOW'

    patients = []
    for r in rows:
        p = dict(r)
        p['risk'] = derive_risk(p['result'], p['confidence'] or 0)
        if p.get('created_at'):
            p['created_at'] = str(p['created_at']).replace(' ', 'T')
        patients.append(p)

    return render_template('analytics.html',
        active_page='analytics',
        doctor=session['doctor'],
        doctor_name=session.get('doctor_name', ''),
        patients_json=json.dumps(patients),
        stats={'total': len(patients)}
    )


@app.route('/evaluation')
def evaluation():
    if 'doctor' not in session:
        return redirect(url_for('login'))
    from datetime import datetime
    return render_template('evaluation.html',
        active_page='evaluation',
        doctor_name=session.get('doctor_name', 'Doctor'),
        now=datetime.now().strftime('%d %b %Y, %I:%M %p')
    )


if __name__ == '__main__':
    app.run(debug=True)